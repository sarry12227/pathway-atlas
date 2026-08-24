"""File-only extraction of one explicitly selected XLSX worksheet."""

from __future__ import annotations

from collections.abc import Mapping
from dataclasses import dataclass
from datetime import date, datetime, time
from io import BytesIO
import math
from pathlib import Path
import re
from typing import Any
import xml.etree.ElementTree as ET
import zipfile

from . import (
    CellStatus,
    ColumnMapping,
    ExtractedRow,
    ExtractedTable,
    MappingError,
    StructuredAdapterError,
    StructuredValidationError,
    coerce_column_mapping,
    derive_coverage,
    read_stable_local_file,
    reject_duplicate_rows,
    resolve_headers,
    validate_monotonicity,
)


class SpreadsheetDependencyError(StructuredAdapterError):
    """Raised when the declared optional spreadsheet dependency is absent."""


def extract_spreadsheet(
    path: str | Path,
    *,
    sheet: str,
    mapping: ColumnMapping | Mapping[str, object],
) -> ExtractedTable:
    if not isinstance(sheet, str) or not sheet or sheet != sheet.strip():
        raise StructuredValidationError("sheet must be a nonempty exact worksheet name")
    column_mapping = coerce_column_mapping(mapping)
    source = read_stable_local_file(path, suffixes=(".xlsx",))
    openpyxl = _load_openpyxl()
    structure = _worksheet_structure(source, sheet)
    formula_book = None
    cached_book = None
    try:
        formula_book = openpyxl.load_workbook(
            BytesIO(source),
            read_only=True,
            data_only=False,
            keep_links=False,
            keep_vba=False,
        )
        cached_book = openpyxl.load_workbook(
            BytesIO(source),
            read_only=True,
            data_only=True,
            keep_links=False,
            keep_vba=False,
        )
        if sheet not in formula_book.sheetnames or sheet not in cached_book.sheetnames:
            raise StructuredValidationError("selected worksheet does not exist")
        formula_sheet = formula_book[sheet]
        cached_sheet = cached_book[sheet]
        formula_rows = list(formula_sheet.iter_rows())
        cached_rows = list(cached_sheet.iter_rows())
        if not formula_rows:
            raise MappingError("selected worksheet has no explicit header row")
        headers = [_header_value(cell.value) for cell in formula_rows[0]]
        positions = resolve_headers(headers, column_mapping)
        width = len(headers)
        extracted: list[ExtractedRow] = []
        explicit_rows = structure.row_cells
        max_rows = max(len(formula_rows), max(explicit_rows, default=1))
        for row_number in range(2, max_rows + 1):
            formula_cells = formula_rows[row_number - 1] if row_number <= len(formula_rows) else ()
            cached_cells = cached_rows[row_number - 1] if row_number <= len(cached_rows) else ()
            if row_number not in explicit_rows and not any(
                getattr(cell, "data_type", None) == "f" or getattr(cell, "value", None) is not None
                for cell in formula_cells
            ):
                continue
            extracted.append(
                _extract_row(
                    formula_cells,
                    cached_cells,
                    row_number,
                    width,
                    positions,
                    column_mapping,
                    structure,
                    formula_sheet.sheet_state != "visible",
                    sheet,
                    openpyxl.utils.get_column_letter,
                )
            )
    except (MappingError, StructuredValidationError):
        raise
    except Exception as error:
        raise StructuredValidationError("XLSX input could not be parsed safely") from error
    finally:
        if formula_book is not None:
            formula_book.close()
        if cached_book is not None:
            cached_book.close()

    reject_duplicate_rows(extracted)
    validate_monotonicity(extracted, column_mapping)
    coverage, coverage_warnings = derive_coverage(extracted, column_mapping)
    warnings: list[str] = []
    if not extracted:
        warnings.append("empty-table")
    if structure.sheet_hidden:
        warnings.append("hidden-sheet")
    if structure.header_hidden:
        warnings.append("hidden-header-row")
    if any(any(status is not CellStatus.EXACT for status in row.cell_status.values()) for row in extracted):
        warnings.append("coverage-excludes-nonexact-rows")
    warnings.extend(coverage_warnings)
    return ExtractedTable(
        table_id=f"sheet:{sheet}",
        caption=None,
        sheet=sheet,
        rows=tuple(extracted),
        coverage=coverage,
        warnings=tuple(warnings),
        extraction_method="xlsx-worksheet",
    )


def _load_openpyxl():
    try:
        import openpyxl
    except (ImportError, ModuleNotFoundError) as error:
        raise SpreadsheetDependencyError(
            "spreadsheet extraction requires openpyxl>=3.1,<4"
        ) from error
    version_text = getattr(openpyxl, "__version__", "")
    try:
        version = tuple(int(part) for part in version_text.split(".")[:2])
    except (TypeError, ValueError) as error:
        raise SpreadsheetDependencyError(
            "spreadsheet extraction requires openpyxl>=3.1,<4"
        ) from error
    if len(version) < 2 or version < (3, 1) or version >= (4, 0):
        raise SpreadsheetDependencyError(
            "spreadsheet extraction requires openpyxl>=3.1,<4"
        )
    return openpyxl


@dataclass(frozen=True)
class _CellRange:
    start_row: int
    end_row: int
    start_column: int
    end_column: int

    def contains(self, row: int, column: int) -> bool:
        return self.start_row <= row <= self.end_row and self.start_column <= column <= self.end_column


@dataclass(frozen=True)
class _ColumnRange:
    start: int
    end: int

    def contains(self, column: int) -> bool:
        return self.start <= column <= self.end


@dataclass(frozen=True)
class _WorksheetStructure:
    merged_ranges: tuple[_CellRange, ...]
    hidden_rows: frozenset[int]
    hidden_column_ranges: tuple[_ColumnRange, ...]
    row_cells: dict[int, frozenset[int]]
    sheet_hidden: bool
    header_hidden: bool

    def is_merged(self, row: int, column: int) -> bool:
        return any(item.contains(row, column) for item in self.merged_ranges)

    def is_hidden_column(self, column: int) -> bool:
        return any(item.contains(column) for item in self.hidden_column_ranges)


_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_PACKAGE_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_CELL_REFERENCE = re.compile(r"^([A-Z]+)([1-9][0-9]*)$")
_MAX_EXCEL_ROWS = 1_048_576
_MAX_EXCEL_COLUMNS = 16_384
_MAX_STRUCTURE_RANGES = 4_096


def _worksheet_structure(source: bytes, sheet: str) -> _WorksheetStructure:
    try:
        with zipfile.ZipFile(BytesIO(source)) as archive:
            workbook = ET.fromstring(archive.read("xl/workbook.xml"))
            relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            relationship_targets = {
                item.attrib["Id"]: item.attrib["Target"]
                for item in relationships.findall(f"{{{_PACKAGE_REL_NS}}}Relationship")
            }
            selected = None
            for item in workbook.findall(f"{{{_MAIN_NS}}}sheets/{{{_MAIN_NS}}}sheet"):
                if item.attrib.get("name") == sheet:
                    selected = item
                    break
            if selected is None:
                raise StructuredValidationError("selected worksheet does not exist")
            relation_id = selected.attrib.get(f"{{{_REL_NS}}}id")
            target = relationship_targets.get(relation_id or "")
            if target is None:
                raise StructuredValidationError("selected worksheet relationship is invalid")
            target = target.replace("\\", "/")
            if target.startswith("/"):
                member = target.lstrip("/")
            elif target.startswith("xl/"):
                member = target
            else:
                member = f"xl/{target.lstrip('./')}"
            worksheet = ET.fromstring(archive.read(member))
    except StructuredValidationError:
        raise
    except (KeyError, OSError, ET.ParseError, zipfile.BadZipFile) as error:
        raise StructuredValidationError("XLSX package structure is invalid") from error

    merge_items = worksheet.findall(f"{{{_MAIN_NS}}}mergeCells/{{{_MAIN_NS}}}mergeCell")
    if len(merge_items) > _MAX_STRUCTURE_RANGES:
        raise StructuredValidationError("worksheet contains too many merged ranges")
    merged_ranges: list[_CellRange] = []
    for item in merge_items:
        reference = item.attrib.get("ref", "")
        try:
            start, end = reference.split(":", 1)
            start_row, start_col = _coordinate(start)
            end_row, end_col = _coordinate(end)
        except (TypeError, ValueError) as error:
            raise StructuredValidationError("worksheet merge range is invalid") from error
        if start_row > end_row or start_col > end_col:
            raise StructuredValidationError("worksheet merge range is reversed")
        merged_ranges.append(_CellRange(start_row, end_row, start_col, end_col))

    hidden_rows: set[int] = set()
    row_cells: dict[int, frozenset[int]] = {}
    for row in worksheet.findall(f"{{{_MAIN_NS}}}sheetData/{{{_MAIN_NS}}}row"):
        try:
            row_number = int(row.attrib["r"])
        except (KeyError, ValueError) as error:
            raise StructuredValidationError("worksheet row identity is invalid") from error
        if not 1 <= row_number <= _MAX_EXCEL_ROWS:
            raise StructuredValidationError("worksheet row identity is outside the Excel grid")
        if row.attrib.get("hidden") in {"1", "true", "True"}:
            hidden_rows.add(row_number)
        cells: set[int] = set()
        for cell in row.findall(f"{{{_MAIN_NS}}}c"):
            reference = cell.attrib.get("r", "")
            parsed_row, parsed_column = _coordinate(reference)
            if parsed_row != row_number:
                raise StructuredValidationError("worksheet cell location is inconsistent")
            cells.add(parsed_column)
        row_cells[row_number] = frozenset(cells)

    hidden_items = [
        column
        for column in worksheet.findall(f"{{{_MAIN_NS}}}cols/{{{_MAIN_NS}}}col")
        if column.attrib.get("hidden") in {"1", "true", "True"}
    ]
    if len(hidden_items) > _MAX_STRUCTURE_RANGES:
        raise StructuredValidationError("worksheet contains too many hidden column ranges")
    hidden_column_ranges: list[_ColumnRange] = []
    for column in hidden_items:
        try:
            lower = int(column.attrib["min"])
            upper = int(column.attrib["max"])
        except (KeyError, ValueError) as error:
            raise StructuredValidationError("worksheet column identity is invalid") from error
        if not 1 <= lower <= upper <= _MAX_EXCEL_COLUMNS:
            raise StructuredValidationError("worksheet column range is outside the Excel grid")
        hidden_column_ranges.append(_ColumnRange(lower, upper))
    return _WorksheetStructure(
        merged_ranges=tuple(merged_ranges),
        hidden_rows=frozenset(hidden_rows),
        hidden_column_ranges=tuple(hidden_column_ranges),
        row_cells=row_cells,
        sheet_hidden=selected.attrib.get("state", "visible") != "visible",
        header_hidden=1 in hidden_rows,
    )


def _coordinate(reference: str) -> tuple[int, int]:
    match = _CELL_REFERENCE.fullmatch(reference)
    if match is None:
        raise StructuredValidationError("worksheet cell reference is invalid")
    letters, row_text = match.groups()
    column = 0
    for character in letters:
        column = column * 26 + ord(character) - ord("A") + 1
    row = int(row_text)
    if row > _MAX_EXCEL_ROWS or column > _MAX_EXCEL_COLUMNS:
        raise StructuredValidationError("worksheet cell reference is outside the Excel grid")
    return row, column


def _header_value(value: object) -> str:
    if not isinstance(value, str) or not value or value != value.strip():
        raise MappingError("worksheet headers must be nonempty exact strings")
    return value


def _extract_row(
    formula_cells: tuple[Any, ...] | list[Any],
    cached_cells: tuple[Any, ...] | list[Any],
    row_number: int,
    width: int,
    positions: dict[str, int],
    mapping: ColumnMapping,
    structure: _WorksheetStructure,
    sheet_hidden: bool,
    sheet_name: str,
    get_column_letter,
) -> ExtractedRow:
    warnings: list[str] = []
    explicit_columns = structure.row_cells.get(row_number, frozenset())
    required_columns = {position + 1 for position in positions.values()}
    truncated = not required_columns.issubset(explicit_columns)
    if truncated:
        warnings.append("truncated-row")
    if row_number in structure.hidden_rows:
        warnings.append("hidden-row")
    structure_uncertain = (
        structure.header_hidden
        or sheet_hidden
        or row_number in structure.hidden_rows
        or truncated
        or any(structure.is_hidden_column(column) for column in required_columns)
    )
    values: dict[str, Any] = {}
    statuses: dict[str, CellStatus] = {}
    for canonical, zero_based in positions.items():
        column_number = zero_based + 1
        formula_cell = formula_cells[zero_based] if zero_based < len(formula_cells) else None
        cached_cell = cached_cells[zero_based] if zero_based < len(cached_cells) else None
        formula_value = getattr(formula_cell, "value", None)
        data_type = getattr(formula_cell, "data_type", None)
        merged = structure.is_merged(row_number, column_number)
        if data_type == "f":
            value = _json_cell_value(getattr(cached_cell, "value", None))
            status = CellStatus.FORMULA
        elif merged:
            value = _json_cell_value(formula_value)
            status = CellStatus.MERGED
        elif data_type == "e":
            value = None
            status = CellStatus.INVALID
        else:
            value, status = _normalize_value(formula_value, mapping.roles.get(canonical), mapping.score_scale)
        if structure_uncertain and status is CellStatus.EXACT:
            status = CellStatus.UNCERTAIN
        values[canonical] = value
        statuses[canonical] = status
        if status is CellStatus.EMPTY:
            warnings.append(f"empty-required-cell:{canonical}")
        elif status is CellStatus.FORMULA:
            warnings.append(f"formula-cell:{canonical}")
        elif status is CellStatus.MERGED:
            warnings.append(f"merged-cell:{canonical}")
        elif status is CellStatus.INVALID:
            warnings.append(f"invalid-cell:{canonical}")
        elif status is CellStatus.UNCERTAIN:
            warnings.append(f"uncertain-cell:{canonical}")
        if merged and status is not CellStatus.MERGED:
            warnings.append(f"merged-cell:{canonical}")
        if structure.is_hidden_column(column_number):
            warnings.append(f"hidden-column:{canonical}")
    nonexact = sum(status is not CellStatus.EXACT for status in statuses.values())
    confidence = max(0.0, 1.0 - (nonexact / max(1, len(statuses))) * 0.5)
    location = f"{sheet_name}!A{row_number}:{get_column_letter(width)}{row_number}"
    return ExtractedRow(values, statuses, location, confidence, tuple(warnings))


def _normalize_value(
    value: object,
    role: str | None,
    score_scale: tuple[int, int] | None,
) -> tuple[Any, CellStatus]:
    if value is None or value == "":
        return None, CellStatus.EMPTY
    if role is None:
        return _json_cell_value(value), CellStatus.EXACT
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise StructuredValidationError("declared numeric cell is invalid")
    if isinstance(value, float) and not math.isfinite(value):
        raise StructuredValidationError("declared numeric cell must be finite")
    if role == "rank":
        if isinstance(value, float) and not value.is_integer():
            raise StructuredValidationError("rank cells must be mathematical integers")
        if value < 1:
            raise StructuredValidationError("rank cells must be positive")
        value = int(value)
    elif isinstance(value, float) and value.is_integer():
        value = int(value)
    if role == "score" and score_scale is not None and not score_scale[0] <= value <= score_scale[1]:
        raise StructuredValidationError("score is outside the declared scale")
    return value, CellStatus.EXACT


def _json_cell_value(value: object) -> Any:
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        if not math.isfinite(value):
            raise StructuredValidationError("worksheet cell must be finite")
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    raise StructuredValidationError("worksheet cell type is not JSON-safe")


__all__ = ["SpreadsheetDependencyError", "extract_spreadsheet"]
